"""file_extractor.py — extract text from PDF, DOCX, XLSX, ZIP, and plain text files."""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

TEXT_EXTS = {
    ".txt", ".md", ".json", ".js", ".jsx", ".ts", ".tsx", ".py",
    ".css", ".html", ".htm", ".yml", ".yaml", ".xml", ".csv",
    ".log", ".ini", ".toml", ".cfg", ".conf", ".env",
    ".bas", ".vbs", ".vba", ".cls", ".frm", ".rsc",
    ".bat", ".cmd", ".ps1", ".sh",
    ".sql", ".rb", ".php", ".java", ".c", ".cpp", ".h", ".hpp",
    ".cs", ".go", ".rs", ".swift", ".kt", ".r", ".m", ".lua",
    ".pl", ".tcl", ".asm",
    ".gitignore", ".dockerfile", ".makefile",
    ".sln", ".csproj", ".pom", ".gradle",
}


def extract_pdf(data: bytes, max_chars: int = 50000) -> str:
    try:
        from app.infrastructure.files.pdf_pro import extract_pdf_smart
        result = extract_pdf_smart(data, max_chars)
        text = result.get("text", "")
        tables = result.get("tables", [])
        if tables:
            table_lines = ["\n\n=== ТАБЛИЦЫ ==="]
            for t in tables[:5]:
                headers = t.get("headers", [])
                rows = t.get("rows", [])
                table_lines.append(f"\nТаблица (стр. {t.get('page', '?')}):")
                if headers:
                    table_lines.append(" | ".join(str(h or "") for h in headers))
                    table_lines.append("-" * 40)
                for row in rows[:20]:
                    table_lines.append(" | ".join(str(c or "") for c in row))
            text += "\n".join(table_lines)
        if result.get("ocr_used"):
            text = f"[OCR распознавание]\n{text}"
        return text[:max_chars]
    except ImportError:
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            parts, total = [], 0
            for page in reader.pages:
                t = page.extract_text() or ""
                if total + len(t) > max_chars:
                    parts.append(t[:max_chars - total])
                    break
                parts.append(t)
                total += len(t)
            return "\n\n".join(parts)
        except ImportError:
            return "[pypdf не установлен: pip install pypdf]"
    except Exception as e:
        return f"[PDF ошибка: {e}]"


def extract_docx(data: bytes, max_chars: int = 30000) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        parts, total = [], 0
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                if total + len(text) > max_chars:
                    break
                parts.append(text)
                total += len(text)
        return "\n".join(parts)
    except ImportError:
        return "[python-docx не установлен: pip install python-docx]"
    except Exception as e:
        return f"[DOCX ошибка: {e}]"


def extract_xlsx(data: bytes, max_chars: int = 30000) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts, total = [], 0
        for sheet in wb.sheetnames[:5]:
            ws = wb[sheet]
            parts.append(f"=== Лист: {sheet} ===")
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(cells)
                if total + len(line) > max_chars:
                    break
                parts.append(line)
                total += len(line)
        wb.close()
        return "\n".join(parts)
    except ImportError:
        return "[openpyxl не установлен: pip install openpyxl]"
    except Exception as e:
        return f"[XLSX ошибка: {e}]"


def extract_zip(data: bytes, max_chars: int = 30000) -> str:
    try:
        parts, total = [], 0
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            parts.append(f"ZIP содержит {len(zf.namelist())} файлов:")
            for name in zf.namelist()[:30]:
                ext = Path(name).suffix.lower()
                size = zf.getinfo(name).file_size
                parts.append(f"  - {name} ({size} байт)")
                if ext in TEXT_EXTS and size < 100_000:
                    try:
                        content = zf.read(name).decode("utf-8", errors="replace")
                        if total + len(content) > max_chars:
                            content = content[:max_chars - total]
                        parts.append(f"\n--- {name} ---\n{content}")
                        total += len(content)
                    except Exception:
                        pass
                if total > max_chars:
                    break
        return "\n".join(parts)
    except Exception as e:
        return f"[ZIP ошибка: {e}]"


def extract_text(data: bytes, max_chars: int = 30000) -> str:
    if not data:
        return ""
    for enc in ("utf-8", "utf-8-sig", "cp1251", "cp866", "latin-1"):
        try:
            text = data.decode(enc)
            if "�" not in text[:500]:
                return text[:max_chars]
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace")[:max_chars]


def extract_any(data: bytes, filename: str, max_chars: int = 50000) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(data, max_chars)
    if ext in (".docx", ".doc"):
        return extract_docx(data, min(max_chars, 30000))
    if ext in (".xlsx", ".xls", ".xlsm"):
        return extract_xlsx(data, min(max_chars, 30000))
    if ext == ".zip":
        return extract_zip(data, min(max_chars, 30000))
    return extract_text(data, min(max_chars, 30000))
