from __future__ import annotations

import hashlib
import io
import sqlite3
from pathlib import Path
from typing import Any

from app.core.config import DATA_DIR, UPLOAD_DIR
from app.infrastructure.db.connection import connect_sqlite

SQLITE_DB = DATA_DIR / "library.db"
UPLOADS_DIR = UPLOAD_DIR

TEXT_EXTS = {
    ".txt",
    ".md",
    ".json",
    ".js",
    ".jsx",
    ".ts",
    ".tsx",
    ".py",
    ".css",
    ".html",
    ".yml",
    ".yaml",
    ".xml",
    ".csv",
    ".log",
    ".ini",
    ".toml",
    ".bas",
    ".vbs",
    ".vba",
    ".cls",
    ".frm",
    ".rsc",
    ".bat",
    ".cmd",
    ".ps1",
    ".sh",
    ".sql",
    ".rb",
    ".php",
    ".java",
    ".c",
    ".cpp",
    ".h",
    ".cs",
    ".go",
    ".rs",
}


def _conn() -> sqlite3.Connection:
    return connect_sqlite(SQLITE_DB, row_factory=sqlite3.Row, journal_mode=None)


def _ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")


def init_library_db() -> None:
    SQLITE_DB.parent.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    conn = _conn()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                size INTEGER DEFAULT 0,
                type TEXT DEFAULT 'unknown',
                preview TEXT DEFAULT '',
                use_in_context INTEGER DEFAULT 1,
                source TEXT DEFAULT 'upload',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        _ensure_column(conn, "files", "stored_path", "stored_path TEXT DEFAULT ''")
        _ensure_column(conn, "files", "sha256", "sha256 TEXT DEFAULT ''")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_lib_ctx ON files(use_in_context)")
        conn.commit()
    finally:
        conn.close()


def safe_disk_name(filename: str, data: bytes) -> str:
    original = Path(filename or "unknown").name
    stem = Path(original).stem or "file"
    suffix = Path(original).suffix
    safe_stem = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in stem)[:80] or "file"
    digest = hashlib.sha256(data).hexdigest()[:12]
    return f"{safe_stem}_{digest}{suffix}"


def extract_preview(filename: str, contents: bytes) -> str:
    ext = Path(filename).suffix.lower()
    preview = ""
    if ext in TEXT_EXTS:
        return contents.decode("utf-8", errors="replace")[:12000]
    if ext == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(contents))
            parts = [(page.extract_text() or "") for page in reader.pages[:20]]
            preview = "\n".join(parts)[:12000]
        except Exception:
            preview = ""
    elif ext in (".docx", ".doc"):
        try:
            from docx import Document

            doc = Document(io.BytesIO(contents))
            preview = "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:12000]
        except Exception:
            preview = ""
    elif ext in (".xlsx", ".xls", ".xlsm"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            parts = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                parts.append(f"=== {sheet} ===")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    parts.append(" | ".join(str(c) if c is not None else "" for c in row))
            preview = "\n".join(parts)[:12000]
            wb.close()
        except Exception:
            preview = ""
    return preview


def read_disk_preview(stored_path: str, max_chars: int) -> str:
    if not stored_path:
        return ""
    try:
        return Path(stored_path).read_text(encoding="utf-8", errors="ignore")[:max_chars]
    except Exception:
        return ""


def list_files() -> dict[str, Any]:
    conn = _conn()
    try:
        rows = conn.execute("SELECT * FROM files ORDER BY created_at DESC, id DESC").fetchall()
    finally:
        conn.close()
    return {"ok": True, "items": [dict(row) for row in rows], "count": len(rows)}


def add_file_contents(
    *,
    filename: str,
    contents: bytes,
    content_type: str | None = None,
    use_in_context: bool = True,
) -> dict[str, Any]:
    filename = filename or "unknown"
    if not contents:
        return {"ok": False, "error": "empty file"}

    disk_name = safe_disk_name(filename, contents)
    disk_path = UPLOADS_DIR / disk_name
    disk_path.write_bytes(contents)
    preview = extract_preview(filename, contents)
    sha256 = hashlib.sha256(contents).hexdigest()

    conn = _conn()
    try:
        cur = conn.execute(
            "INSERT INTO files (name, size, type, preview, use_in_context, source, stored_path, sha256) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                filename,
                len(contents),
                content_type or Path(filename).suffix.lower() or "unknown",
                preview,
                1 if use_in_context else 0,
                "upload",
                str(disk_path),
                sha256,
            ),
        )
        file_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    return {
        "ok": True,
        "id": file_id,
        "name": filename,
        "preview_len": len(preview),
        "stored_path": str(disk_path),
    }


def toggle_context(file_id: int, *, enabled: bool = True) -> dict[str, Any]:
    conn = _conn()
    try:
        conn.execute("UPDATE files SET use_in_context = ? WHERE id = ?", (1 if enabled else 0, file_id))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "id": file_id, "use_in_context": enabled}


def delete_file(file_id: int) -> dict[str, Any]:
    conn = _conn()
    try:
        row = conn.execute("SELECT stored_path FROM files WHERE id = ?", (file_id,)).fetchone()
        conn.execute("DELETE FROM files WHERE id = ?", (file_id,))
        conn.commit()
    finally:
        conn.close()

    if row and row["stored_path"]:
        try:
            path = Path(row["stored_path"])
            if path.exists() and path.is_file():
                path.unlink()
        except Exception:
            pass
    return {"ok": True, "deleted_id": file_id}


def search_files(query: str = "") -> dict[str, Any]:
    q = f"%{query.strip()}%"
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT * FROM files WHERE name LIKE ? OR preview LIKE ? ORDER BY created_at DESC, id DESC LIMIT 20",
            (q, q),
        ).fetchall()
    finally:
        conn.close()
    return {"ok": True, "items": [dict(row) for row in rows], "count": len(rows)}


def get_context_files() -> dict[str, Any]:
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT id, name, preview, stored_path FROM files WHERE use_in_context = 1 AND preview != '' ORDER BY created_at DESC, id DESC LIMIT 10"
        ).fetchall()
    finally:
        conn.close()
    return {"ok": True, "items": [dict(row) for row in rows], "count": len(rows)}


def list_library_files() -> dict[str, Any]:
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT id, name, size, use_in_context, stored_path, type, source, created_at FROM files ORDER BY created_at DESC, id DESC"
        ).fetchall()
    finally:
        conn.close()

    files = []
    for row in rows:
        files.append(
            {
                "id": row["id"],
                "name": row["name"],
                "size": row["size"] or 0,
                "active": bool(row["use_in_context"]),
                "path": row["stored_path"] or str(UPLOADS_DIR / row["name"]),
                "suffix": Path(row["name"]).suffix.lower(),
                "type": row["type"] or "unknown",
                "source": row["source"] or "upload",
                "created_at": row["created_at"],
            }
        )
    return {"ok": True, "files": files, "count": len(files)}


def set_library_active(filename: str, active: bool) -> dict[str, Any]:
    conn = _conn()
    try:
        row = conn.execute("SELECT id, name FROM files WHERE name = ? ORDER BY id DESC LIMIT 1", (filename,)).fetchone()
        if not row:
            return {"ok": False, "error": f"Файл не найден: {filename}"}
        conn.execute("UPDATE files SET use_in_context = ? WHERE id = ?", (1 if active else 0, row["id"]))
        conn.commit()
        return {"ok": True, "filename": filename, "active": bool(active)}
    finally:
        conn.close()


def delete_library_file(filename: str) -> dict[str, Any]:
    conn = _conn()
    try:
        row = conn.execute("SELECT id, stored_path FROM files WHERE name = ? ORDER BY id DESC LIMIT 1", (filename,)).fetchone()
        if not row:
            return {"ok": False, "error": f"Файл не найден: {filename}"}
        conn.execute("DELETE FROM files WHERE id = ?", (row["id"],))
        conn.commit()
    finally:
        conn.close()

    stored_path = row["stored_path"] or ""
    if stored_path:
        try:
            path = Path(stored_path)
            if path.exists() and path.is_file():
                path.unlink()
        except Exception:
            pass
    return {"ok": True, "filename": filename}


def build_library_context(max_files: int = 3, max_chars_per_file: int = 4000) -> dict[str, Any]:
    conn = _conn()
    try:
        rows = conn.execute(
            "SELECT name, preview, stored_path FROM files WHERE use_in_context = 1 ORDER BY created_at DESC, id DESC LIMIT ?",
            (max_files,),
        ).fetchall()
    finally:
        conn.close()

    context_parts = []
    used_files = []
    active_count = 0
    for row in rows:
        active_count += 1
        suffix = Path(row["name"]).suffix.lower()
        content = (row["preview"] or "")[:max_chars_per_file]
        if not content and suffix in TEXT_EXTS:
            content = read_disk_preview(row["stored_path"] or "", max_chars_per_file)
        if not content.strip():
            continue
        used_files.append(row["name"])
        context_parts.append(f"===== FILE: {row['name']} =====\n{content}")

    return {
        "ok": True,
        "used_files": used_files,
        "active_count": active_count,
        "context": "\n\n".join(context_parts),
    }


init_library_db()
