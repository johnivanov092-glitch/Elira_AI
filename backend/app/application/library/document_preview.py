"""Document preview extraction and safe filename generation."""
from __future__ import annotations

import hashlib
from pathlib import Path

_TEXT_EXTS = {
    ".txt", ".md", ".json", ".js", ".jsx", ".ts", ".tsx", ".py", ".css",
    ".html", ".yml", ".yaml", ".xml", ".csv", ".log", ".ini", ".toml",
    ".bas", ".vbs", ".vba", ".cls", ".frm", ".rsc", ".bat", ".cmd",
    ".ps1", ".sh", ".sql", ".rb", ".php", ".java", ".c", ".cpp", ".h", ".cs",
    ".go", ".rs",
}


def safe_disk_name(filename: str, data: bytes) -> str:
    original = Path(filename or "unknown").name
    stem = Path(original).stem or "file"
    suffix = Path(original).suffix
    safe_stem = "".join(ch if ch.isalnum() or ch in ("-", "_", ".") else "_" for ch in stem)[:80] or "file"
    digest = hashlib.sha256(data).hexdigest()[:12]
    return f"{safe_stem}_{digest}{suffix}"


def extract_preview(filename: str, contents: bytes) -> str:
    ext = Path(filename).suffix.lower()

    if ext in _TEXT_EXTS:
        try:
            return contents.decode("utf-8", errors="replace")[:12000]
        except Exception:
            return ""

    if ext == ".pdf":
        try:
            import io
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(contents))
            parts = [page.extract_text() or "" for page in reader.pages[:20]]
            return "\n".join(parts)[:12000]
        except Exception:
            return ""

    if ext in (".docx", ".doc"):
        try:
            import io
            from docx import Document
            doc = Document(io.BytesIO(contents))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())[:12000]
        except Exception:
            return ""

    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                parts.append(f"=== {sheet} ===")
                for row in ws.iter_rows(max_row=100, values_only=True):
                    parts.append(" | ".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(parts)[:12000]
        except Exception:
            return ""

    return ""
